from asyncio.windows_events import NULL
from datetime import datetime
from turtle import heading
import requests
import PyPDF2
import json
import io
import re
import openpyxl


"""
////////////////////////////////////////////////////////////////////////////////
///////////////////////////////////// FUNCTION DEFINITIONS /////////////////////
////////////////////////////////////////////////////////////////////////////////
"""
# This function will return a list of all the id's for the pdfs that we want to search through
def accessID(startYear, endYear, chamberArray):
    idList = []
    for year in range(startYear, endYear + 1):
        response = requests.get("https://api.parliament.nsw.gov.au/api/hansard/search/year/" + str(year))
        if response.status_code == 200:
            response = json.loads(response.text)
        elif (response.status_code == 403 or response.status_code == 404):
            print("Error accessing resource. Exiting...")
            exit()

        # for each item in the response, get its events, and in the events array, loop through and obtain the id for each event
        # append chamber to the chamberArray to keep track of the chamber for use later on
        for dateItem in response:
            for eventItems in dateItem['Events']:
                chamberArray.append(eventItems['Chamber'])
                idList.append(eventItems['PdfDocId'])
        pass
    return idList

# This function will return the pdf content for a given id
def getPDF(id):
    print("Accessing id", id, "...")
    response = requests.get("https://api.parliament.nsw.gov.au/api/hansard/search/daily/pdf/" + id)
    if response.status_code == 200:
        print("Success!")
        return response.content
    else:
        #For handling invalid HANSARD documents
        print("failure")
        return 'invalid'

# This function will search through the pdf content for a given search term
#TODO: use a dictionary to store information easier. One can clearly see
# that I am trained in C. 
def search_pdf(pdf_content, search_terms):
    pdf_file = io.BytesIO(pdf_content)
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    
    total_occurrences = [0 for _ in range (len(search_terms))]

    for page_num in range(len(pdf_reader.pages)):
        page = pdf_reader.pages[page_num]
        text = page.extract_text()
        
        # Remove line breaks and other whitespace characters
        cleaned_text = re.sub(r'\n', ' ', text)
        
        for i, term in enumerate(searchTerms):
            pattern = re.compile(r'\b' + re.escape(term.lower()) + r'\b')
            occurrences_on_page = len(pattern.findall(cleaned_text.lower()))
            total_occurrences[i] += occurrences_on_page
        
    for i, term in enumerate(searchTerms):
        print(f"Total occurances of {term}: {total_occurrences[i]}")
    return total_occurrences


# This function will create an excel file with the given headings
def createExcelFile(filePath, headings):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for col_num, heading in enumerate(headings, 1):
        sheet.cell(row=1, column=col_num, value=heading)
    workbook.save(filePath)

# This function will add data to an existing excel file
def add_data_to_excel(file_path, data):
    # Open the existing Excel workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the default sheet (index 0)
    sheet = workbook.active

    # Find the next empty row
    next_row = sheet.max_row + 1

    # Add data to the sheet
    for col_num, value in enumerate(data, 1):
        sheet.cell(row=next_row, column=col_num, value=value)

    # Save the changes to the workbook
    workbook.save(file_path)

"""
////////////////////////////////////////////////////////////////////////////////
///////////////////////// MAIN FUNCTION ////////////////////////////////////////
////////////////////////////////////////////////////////////////////////////////
"""
year_string = input("Enter the start year and the end year separated by a space (1991 and beyond): ")
year_string = [int(year) for year in year_string.split()]

if (year_string[0] > year_string[1] or year_string[0] <= 1991 or year_string[1] > datetime.now().year):
    print("Year input incorrect or invalid. Exiting...")
    exit()

search_string = input("Enter your search terms, separated by commas: ")
searchTerms = [term.strip() for term in search_string.split(',')]

chambers = []
pdf_ID = accessID(year_string[0], year_string[1], chambers)

headings = [
    "PDF ID", 
    "Chamber"
]

for term in searchTerms:
    headings.append(f'"{term}" mentioned?')
for term in searchTerms:
    headings.append(f'# of times "{term}" mentioned?')

file_path = 'output.xlsx'
createExcelFile(file_path, headings)

for PDFid in pdf_ID:
    response = getPDF(PDFid)
    if (response == 'invalid'):
        continue
    dataToAdd = []
    dataToAdd.append(PDFid)
    dataToAdd.append(chambers[pdf_ID.index(PDFid)])


    searchResults = []
    searchResults = search_pdf(response, searchTerms)

    for results in searchResults:
        if results > 0:
            dataToAdd.append("Yes")
        else:
            dataToAdd.append("No")

    dataToAdd.extend(searchResults)
    add_data_to_excel(file_path, dataToAdd)
    print("\n")